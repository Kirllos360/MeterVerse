"""
Phase F2: BTU Invoice Formula Reconstruction
Verifies Total = Consumption x Rate against ALL BTU invoice files across months/properties.

Also checks: taxes=0, fees=0, customer_service=0, admin_fees=0.
Reports any anomalies.
"""
import openpyxl
import os
import json
from collections import defaultdict

GC_INVOICES = [
    r"D:\Operation\Months\02-2025\Golf Central 02-2025\Golf Central Mall BTU Invoices 02-2025.xlsx",
    r"D:\Operation\Months\03-2025\Done\golf central\Golf Central Mall BTU Invoices 03-2025.xlsx",
    r"D:\Operation\Months\04-2025\Done\GC\Golf Central Mall BTU Invoices 04-2025.xlsx",
    r"D:\Operation\Months\05-2025\GC\Golf Central Mall BTU Invoices 05-2025.xlsx",
    r"D:\Operation\Months\06-2025\Done\GC\Golf Central Mall BTU Invoices 06-2025.xlsx",
    r"D:\Operation\Months\07-2025\Done\GC\Golf Central Mall BTU Invoices 07-2025.xlsx",
    r"D:\Operation\Months\08-2025\Done\GC\Golf Central Mall BTU Invoices 08-2025.xlsx",
    r"D:\Operation\Months\09-2025\DONE\GC\Golf Central Mall BTU Invoices 09-2025.xlsx",
    r"D:\Operation\Months\10-2025\Done\GC\Golf Central Mall BTU Invoices 10-2025.xlsx",
    r"D:\Operation\Months\11-2025\Done\GC\Golf Central Mall BTU Invoices 11-2025.xlsx",
    r"D:\Operation\Months\12-2025\Done\GC\Golf Central Mall BTU Invoices 12-2025.xlsx",
]

PC_INVOICES = [
    r"D:\Operation\Months\02-2025\PC Report 02-2025\Palm Central BTU Invoices 02-2025.xlsx",
    r"D:\Operation\Months\03-2025\Done\PC-03-2025\Palm Central BTU Invoices 03-2025.xlsx",
    r"D:\Operation\Months\04-2025\Done\PC\Palm Central BTU Invoices 04-2025.xlsx",
    r"D:\Operation\Months\05-2025\Done\PC\Palm Central BTU Invoices 05-2025.xlsx",
    r"D:\Operation\Months\06-2025\Done\PC\Palm Central BTU Invoices 06-2025.xlsx",
    r"D:\Operation\Months\07-2025\Done\PC\Palm Central BTU Invoices 07-2025.xlsx",
    r"D:\Operation\Months\08-2025\Done\pc\Palm Central BTU Invoices 08-2025.xlsx",
    r"D:\Operation\Months\09-2025\DONE\PC\Palm Central BTU Invoices 09-2025.xlsx",
    r"D:\Operation\Months\10-2025\Done\PC\Palm Central BTU Invoices 10-2025.xlsx",
    r"D:\Operation\Months\11-2025\Done\PC\Palm Central BTU Invoices 11-2025.xlsx",
    r"D:\Operation\Months\12-2025\Done\PC\Palm Central BTU Invoices 12-2025.xlsx",
]

IZAR_INVOICES = [
    r"D:\Operation\Months\03-2025\IZAR 03-2025\IZAR Mall BTU Invoices 03-2025.xlsx",
    r"D:\Operation\Months\05-2025\Done\Izar\04-2025\Izar Mall BTU Invoices 04-2025.xlsx",
    r"D:\Operation\Months\05-2025\Done\Izar\05-2025\Izar Mall BTU Invoices 05-2025.xlsx",
    r"D:\Operation\Months\06-2025\Done\Izar\Izar Mall BTU Invoices 06-2025.xlsx",
    r"D:\Operation\Months\07-2025\Done\Izar\Izar Mall BTU Invoices 07-2025.xlsx",
    r"D:\Operation\Months\08-2025\Done\IZAR\Izar Mall BTU Invoices 08-2025.xlsx",
    r"D:\Operation\Months\09-2025\DONE\IZAR\Izar Mall BTU Invoices 09-2025.xlsx",
    r"D:\Operation\Months\10-2025\Done\IZAR\Izar Mall BTU Invoices 10-2025.xlsx",
    r"D:\Operation\Months\11-2025\Done\IZAR\Izar Mall BTU Invoices 11-2025.xlsx",
]

CORRECTION_INVOICES = [
    r"D:\Operation\Months\12-2025\Done\GC\Golf Central Mall BTU Correction Invoices 12-2025.xlsx",
]

PREVIOUS_INVOICES = [
    r"D:\Operation\Months\02-2025\Golf Central 02-2025\Golf Central Mall Previous BTU Invoices.xlsx",
    r"D:\Operation\Months\06-2025\Done\GC\Golf Central Mall Pervious BTU Invoices.xlsx",
    r"D:\Operation\Months\09-2025\DONE\GC\Golf Central Mall Pervious BTU Invoices.xlsx",
    r"D:\Operation\Months\11-2025\Done\GC\Golf Central Mall BTU Previous Invoices.xlsx",
    r"D:\Operation\Months\07-2025\Done\PC\Palm Central BTU Previous Invoices.xlsx",
    r"D:\Operation\Months\11-2025\Done\PC\Palm Central BTU Previous Invoices.xlsx",
]


def extract_month_from_path(path):
    parts = path.split(os.sep)
    for p in parts:
        if any(m in p for m in ['01-2025','02-2025','03-2025','04-2025','05-2025','06-2025',
                                 '07-2025','08-2025','09-2025','10-2025','11-2025','12-2025',
                                 '01-2026','02-2026','03-2026','04-2026']):
            # Extract YYYY-MM pattern
            import re
            m = re.search(r'(\d{2}-\d{4})', p)
            if m:
                return m.group(1)
    return 'unknown'


def read_btu_invoice_xlsx(path):
    """Read a BTU invoice XLSX and return list of row dicts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.column] = str(cell.value).strip()
    
    col_map = {}
    for col_idx, hdr in headers.items():
        hl = hdr.lower()
        if 'bill' in hl or 'number' in hl:
            col_map['bill'] = col_idx
        elif 'customer' in hl or 'name' in hl or 'عميل' in hl:
            col_map['customer'] = col_idx
        elif 'meter' in hl or 'serial' in hl or 'عداد' in hl:
            col_map['meter'] = col_idx
        elif 'type' in hl:
            col_map['type'] = col_idx
        elif 'unit' in hl or 'unit no' in hl or 'unit number' in hl or 'unit umber' in hl:
            col_map['unit'] = col_idx
        elif 'consumption' in hl or 'kwh' in hl or 'استهلاك' in hl:
            col_map['consumption'] = col_idx
        elif 'tax' in hl or 'ضريبة' in hl:
            col_map['tax'] = col_idx
        elif 'fee' in hl or 'رسوم' in hl:
            col_map['fees'] = col_idx
        elif 'service' in hl or 'خدمة' in hl:
            col_map['service'] = col_idx
        elif 'admin' in hl:
            col_map['admin'] = col_idx
        elif 'total' in hl or 'amount' in hl or 'اجمالى' in hl or 'اجمالي' in hl or 'الإجمالي' in hl or 'total amount' in hl:
            col_map['total'] = col_idx
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        vals = {}
        for key, col_idx in col_map.items():
            cell = row[col_idx - 1]
            vals[key] = cell.value
        
        bill = vals.get('bill')
        if bill is None or bill == '' or bill == 'Bill Number':
            continue
        
        r = {
            'row': row_idx,
            'bill': bill,
            'customer': vals.get('customer', ''),
            'meter': vals.get('meter', ''),
            'unit': vals.get('unit', ''),
            'type': vals.get('type', ''),
            'consumption': vals.get('consumption', 0),
            'tax': vals.get('tax', 0),
            'fees': vals.get('fees', 0),
            'service': vals.get('service', 0),
            'admin': vals.get('admin', 0),
            'total': vals.get('total', 0),
        }
        
        for k in ['consumption', 'tax', 'fees', 'service', 'admin', 'total']:
            v = r[k]
            if v is None:
                r[k] = 0.0
            elif isinstance(v, str):
                v = v.replace(',', '').strip()
                try:
                    r[k] = float(v)
                except:
                    r[k] = 0.0
            else:
                try:
                    r[k] = float(v)
                except:
                    r[k] = 0.0
        
        rows.append(r)
    
    wb.close()
    return rows


results = {
    'files_processed': 0,
    'total_rows': 0,
    'match_count': 0,
    'mismatch_count': 0,
    'anomalies': [],
    'by_property': defaultdict(lambda: {'rows': 0, 'match': 0, 'mismatch': 0}),
    'by_month': defaultdict(lambda: {'rows': 0, 'match': 0, 'mismatch': 0}),
    'rate_counts': defaultdict(int),
}

def process_file(path, label):
    month_label = extract_month_from_path(path)
    if not os.path.exists(path):
        results['anomalies'].append({'file': path, 'label': label, 'error': 'NOT_FOUND'})
        print(f"  NOT FOUND: {path}")
        return
    
    rows = read_btu_invoice_xlsx(path)
    results['files_processed'] += 1
    results['total_rows'] += len(rows)
    
    print(f"  {label} ({month_label}): {len(rows)} rows")
    
    for r in rows:
        cons = r['consumption']
        total = r['total']
        tax = r['tax']
        fees = r['fees']
        service = r['service']
        admin = r['admin']
        
        if cons == 0 and total == 0:
            results['match_count'] += 1
            results['by_property'][label]['match'] += 1
            results['by_month'][month_label]['match'] += 1
            continue
        
        rate = round(total / cons, 6) if cons > 0 else 0
        expected = round(cons * 3.0, 2)
        match = abs(total - expected) < 0.015
        
        if match:
            results['match_count'] += 1
            results['by_property'][label]['match'] += 1
            results['by_month'][month_label]['match'] += 1
        else:
            # Check if it's a different rate
            alt_rate = round(total / cons, 6) if cons > 0 else 0
            alt_expected = round(cons * alt_rate, 2)
            alt_match = abs(total - alt_expected) < 0.015
            
            if alt_match and alt_rate != 3.0:
                results['anomalies'].append({
                    'file': path, 'label': label,
                    'row': r['row'], 'customer': r['customer'],
                    'consumption': cons, 'total': total,
                    'expected_3': expected,
                    'actual_rate': alt_rate,
                    'type': 'DIFFERENT_RATE'
                })
                results['mismatch_count'] += 1
                results['by_property'][label]['mismatch'] += 1
                results['by_month'][month_label]['mismatch'] += 1
                print(f"    Row {r['row']}: {r['customer']} - cons={cons} total={total} -> rate={alt_rate} (!=3.0)")
            else:
                results['anomalies'].append({
                    'file': path, 'label': label,
                    'row': r['row'], 'customer': r['customer'],
                    'consumption': cons, 'total': total,
                    'expected_3': expected,
                    'type': 'FORMULA_MISMATCH'
                })
                results['mismatch_count'] += 1
                results['by_property'][label]['mismatch'] += 1
                results['by_month'][month_label]['mismatch'] += 1
                print(f"    Row {r['row']}: {r['customer']} - cons={cons} total={total} expected={expected} -> MISMATCH")
        
        results['rate_counts'][round(rate, 6)] += 1
        
        # Check non-zero taxes/fees
        if abs(tax) > 0.01:
            results['anomalies'].append({
                'file': path, 'label': label,
                'row': r['row'], 'customer': r['customer'],
                'field': 'tax', 'value': tax,
                'type': 'NONZERO_TAX'
            })
        if abs(fees) > 0.01:
            results['anomalies'].append({
                'file': path, 'label': label,
                'row': r['row'], 'customer': r['customer'],
                'field': 'fees', 'value': fees,
                'type': 'NONZERO_FEES'
            })
        if abs(service) > 0.01:
            results['anomalies'].append({
                'file': path, 'label': label,
                'row': r['row'], 'customer': r['customer'],
                'field': 'service', 'value': service,
                'type': 'NONZERO_SERVICE'
            })
        if abs(admin) > 0.01:
            results['anomalies'].append({
                'file': path, 'label': label,
                'row': r['row'], 'customer': r['customer'],
                'field': 'admin', 'value': admin,
                'type': 'NONZERO_ADMIN'
            })


print("=" * 80)
print("PHASE F2: BTU INVOICE FORMULA RECONSTRUCTION")
print("Formula hypothesis: Total = Consumption x 3.0, Taxes=Fees=Service=Admin=0")
print("=" * 80)

print("\n--- Golf Central Mall BTU Invoices ---")
for f in GC_INVOICES:
    process_file(f, 'GC')

print("\n--- Palm Central BTU Invoices ---")
for f in PC_INVOICES:
    process_file(f, 'PC')

print("\n--- IZAR Mall BTU Invoices ---")
for f in IZAR_INVOICES:
    process_file(f, 'IZAR')

print("\n--- Correction Invoices ---")
for f in CORRECTION_INVOICES:
    process_file(f, 'CORRECTION')

print("\n--- Previous BTU Invoices (for reference) ---")
for f in PREVIOUS_INVOICES:
    process_file(f, 'PREVIOUS')

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"Files processed: {results['files_processed']}")
print(f"Total rows:      {results['total_rows']}")
print(f"Formula MATCH:   {results['match_count']}")
print(f"Formula MISMATCH:{results['mismatch_count']}")
match_pct = 100 * results['match_count'] / results['total_rows'] if results['total_rows'] > 0 else 0
print(f"Match %:         {match_pct:.1f}%")
print(f"Total anomalies: {len(results['anomalies'])}")
print(f"\nRate distribution:")
for rate, count in sorted(results['rate_counts'].items()):
    print(f"  Rate {rate}: {count} rows")

print(f"\nBy property:")
for prop, data in sorted(results['by_property'].items()):
    pct = 100 * data['match'] / data['rows'] if data['rows'] > 0 else 0
    print(f"  {prop}: {data['match']}/{data['rows']} match ({pct:.1f}%)")

print(f"\nBy month:")
for month in sorted(results['by_month'].keys()):
    data = results['by_month'][month]
    pct = 100 * data['match'] / data['rows'] if data['rows'] > 0 else 0
    print(f"  {month}: {data['match']}/{data['rows']} match ({pct:.1f}%)")

# Save report
report = {
    'files_processed': results['files_processed'],
    'total_rows': results['total_rows'],
    'match_count': results['match_count'],
    'mismatch_count': results['mismatch_count'],
    'match_percentage': round(match_pct, 1),
    'by_property': {k: dict(v) for k, v in results['by_property'].items()},
    'by_month': {k: dict(v) for k, v in results['by_month'].items()},
    'rate_distribution': {str(k): v for k, v in results['rate_counts'].items()},
    'anomalies': results['anomalies'][:50],  # limit
}

os.makedirs('reports', exist_ok=True)
with open('reports/f2-btu-formula-reconstruction.json', 'w', encoding='utf-8') as f:
    json.dump(report, f, ensure_ascii=False, indent=2)

print(f"\nDetailed report saved to reports/f2-btu-formula-reconstruction.json")
